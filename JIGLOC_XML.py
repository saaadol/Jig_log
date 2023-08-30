from tkinter.filedialog import askdirectory

import openpyxl
import pandas as pd
import xml.etree.ElementTree as ET
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
from ttkthemes import ThemedStyle
current_directory = os.getcwd()

def switch_letters_and_numbers(value):
    # Switch letters and numbers as before
    numeric_part = ''.join(filter(str.isdigit, value))
    letter_part = ''.join(filter(str.isalpha, value))
    return letter_part + numeric_part

def determine_type(fin_value):
    fin_value_str = str(fin_value)
    if 'SP' in fin_value or 'VP' in fin_value:
        return 'Passthrough'
    else:
        return 'Connector'

def get_name_from_matching_column(sheet, title):
    # Find the corresponding value in the 'matching' column (3rd column)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == title:
            return str(row[3])
    return None

def convert_to_xml(file, output_directory):

    # Extract information from the input file name
    file_name = os.path.basename(file)
    file_parts = file_name.split('.')[0].split('-')
    title = file_parts[0] + '-' + file_parts[1] + '-' + file_parts[2]
    version = file_parts[2]
    name1 = file_parts[0]
    solution = file_parts[1]

    # Assuming the data is in the first sheet
    #backup
    #matching_file = browse_matching_file_function()
    #workbook1 = openpyxl.load_workbook(matching_file)
    
    sheet1 = workbook1.active
    # Find the name value from the 'matching' column in the XLSX file
    name_from_matching = get_name_from_matching_column(sheet1, title)
    if name_from_matching is not None:
        name = name_from_matching
    else:
        name = name1

    workbook = openpyxl.load_workbook(file)
    sheet = workbook['FINLoc']
    # Create the root element of the XML
    root = ET.Element('IGE-XAO-JIGBOARD', XSDVersion="4.00")
    # Create the header element
    header = ET.SubElement(root, 'Header', Harness=name, Version='*')
    # Create the jig board element
    jig_board = ET.SubElement(root, 'JigBoard', Name=name1, Solution=solution, Version=version,
                              JigboardOrder="001")
    # Create the devices list element
    devices_list = ET.SubElement(jig_board, 'DevicesList')

    for row in sheet.iter_rows(min_row=6, values_only=True):
        fin_value = row[0]
        localisation_value = row[1]


        # Get the values from the respective columns
        tag = fin_value
        localisation = switch_letters_and_numbers(localisation_value)
        type = determine_type(fin_value)
        #location = str(sheet.cell(row=row, column=2).value)
        #device_type = str(sheet.cell(row=row, column=3).value)

        # Break the loop if the tag is empty
        if tag is None:
            break

        # Create the connective device element
        connective_device = ET.SubElement(devices_list, 'ConnectiveDevice', Tag=tag, Location=localisation,
                                          Type=type)

    # Create the user attribute element
    user_attribute = ET.SubElement(root, 'UserAttribute', AttributeName='CustomerSolution',
                                   AttributeValue=f"{name1}-{solution}-{version}")


    # Generate the XML file name
    xml_file_name = os.path.splitext(os.path.basename(file))[0] + '.xml'
    xml_file_path = os.path.join(output_directory, xml_file_name)
    # Create the XML tree from the root element
    tree = ET.ElementTree(root)
    # Write the XML tree to the output file
    tree.write(xml_file_path, encoding='utf-8', xml_declaration=True)


def browse_files():
    path = filedialog.askdirectory(title="Select Input directory")
    if path:
        output_dir = filedialog.askdirectory(title="Select Output directory")
        if output_dir:
            browse_matching_file_function()
            process_files(path, output_dir)

def process_files(input_path, output_path):
    for filename in os.listdir(input_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(input_path, filename)
            convert_to_xml(file_path, output_path)
    messagebox.showinfo("Message", 'XML files have been created successfully')
    window.destroy()


def browse_matching_file_function():
    global workbook1
    file = filedialog.askopenfilename(title="Select Matching File",
                                            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    if file:
        workbook1 = openpyxl.load_workbook(f'{file}')
    else:
        exit()


# Create the GUI window
window = tk.Tk()
window.title('XML Generator')
window.geometry('281x130')
window.resizable(width=False, height=False)
window.iconbitmap(f"{current_directory}\\favicon.ico")
# Create a themed style for the window
style = ThemedStyle(window)
style.set_theme("elegance")  # Set the theme to "Arc"

# Configure the style
style.configure("TLabel", forground="", font=("Helvetica", 9))
style.configure("TButton", forground="", font=("Helvetica", 9))
style.configure("TEntry", forground="", font=("Helvetica", 9))

# Load the background image
background_image_path = f"{current_directory}\\bg.jpg"  # Replace with the actual path to your background image
if os.path.exists(background_image_path):
    background_image = Image.open(background_image_path)
   # background_image = background_image.resize((700, 700))  # Adjust the size of the image to match the window
    background_photo = ImageTk.PhotoImage(background_image)

    # Create a canvas and display the background image
    canvas = tk.Canvas(window, width=1200, height=390)  # Adjust the canvas size to match the window
    canvas.create_image(0, 0, anchor=tk.NW, image=background_photo)
    canvas.pack()

# Create the Browse button
style = ThemedStyle(window)
style.set_theme("elegance") 
browse_button = ttk.Button(window, text='Browse Files', command=browse_files)
browse_button.place(relx=0.80, rely=0.42, anchor=tk.CENTER)


# Create the Process Files button
#process_button = tk.Button(window, text='Process Files', command=process_file)
#process_button.place(relx=0.485, rely=0.5, anchor=tk.CENTER)

# Run the GUI event loop
window.mainloop()